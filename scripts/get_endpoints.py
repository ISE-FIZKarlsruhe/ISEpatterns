import httpx
from io import BytesIO
from rdflib import Graph, URIRef, Literal, Namespace
from openpyxl import load_workbook

u = "https://docs.google.com/spreadsheets/d/1HVxe9DoKtJTfJHl0l_NQKH-3CVxv-LTHa6xMHoYBcFk/export?format=xlsx"

r = httpx.get(u, follow_redirects=True)
if r.status_code == 200:
    data = r.content

    workbook = load_workbook(filename=BytesIO(data))
    sheet = workbook.active

    g = Graph()
    for row in sheet.iter_rows(max_col=1, values_only=True):
        if row[0] is None or not row[0].startswith("http"):
            continue
        object = URIRef(str(row[0]))
        predicate = URIRef("https://schema.org/URL")
        subject = URIRef("http://designpatterns.ise.fiz-karlsruhe.de/endpoints/")
        g.add((subject, predicate, object))

    ttl = g.serialize(format="turtle")
    open("../data/endpoints.ttl", "w").write(ttl)
